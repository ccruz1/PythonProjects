
# importing pandas as pd
from datetime import datetime 
import pandas as pd
import openpyxl 
from openpyxl import load_workbook
import os
  
x = 0
start_time = datetime.now() 

##Ask the year needed
year_Needed = input("Ingresa el año:") 

##Create excel file to export
excel_Name = year_Needed +".xlsx"
wb = openpyxl.Workbook()
wb.save(excel_Name)

##List files per year
##Change the year as needed
path_Default = "DTG/"+ year_Needed +"/"
list_Files = os.listdir(path_Default)

##Open Excel for wrting
book = load_workbook(excel_Name)
writer = pd.ExcelWriter(excel_Name, engine = 'openpyxl')
writer.book = book

##Create the sheets in excel file
for files in list_Files:
    print("Escribiendo...")
    df = pd.DataFrame(pd.read_excel(path_Default+files))
    df = df[df['sector_economico_4'].astype('str').str.contains('8701')]
    df.to_excel(writer, sheet_name = files, index = False)
    x+=1
    print(str(x) + " Hoja(s) Escrita(s)")

##Closes excel write file
writer.close()

##Gets End Script Time
end_time = datetime.now()
print('Duration: {}'.format(end_time - start_time))





